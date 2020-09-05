from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()

# Create workbook instance

wb = Workbook()

# Load existing workbook
wb = load_workbook('검교정관리대장.xlsx')

#Create active worksheet
ws =wb.active


# Create variable for Column A
column_a = ws['A']
column_b = ws['B']


def get_a():
    for cell in column_a:
        print(cell.value)
        


def get_b():
    for cell in column_b:
        print(cell.value)


btn_a = Button(root, text="Get column A", command=get_a)
btn_a.pack()



print(column_a, column_b)



root.mainloop()